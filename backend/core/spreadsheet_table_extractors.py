from __future__ import annotations

import subprocess
from pathlib import Path
from typing import Dict, List, Tuple

from core.conversion_common import ConversionError, find_libreoffice as _find_libreoffice
from core.spreadsheet_formatting import (
    _build_html_table,
    _build_markdown_table,
    _cell_value_to_str,
    _normalize_excel_text_value,
    _trim_empty_trailing,
)

def _is_header_row(ws, row_index: int, max_col: int) -> bool:
    """1행에 font.bold 또는 fill.fgColor가 있으면 헤더로 간주"""
    for col in range(1, max_col + 1):
        try:
            cell = ws.cell(row=row_index, column=col)
            if cell.font and cell.font.bold:
                return True
            if cell.fill and cell.fill.fill_type not in (None, "none"):
                fg = cell.fill.fgColor
                if fg and fg.type != "none":
                    # 흰색·검정·투명이 아닌 배경색이면 헤더로 간주
                    if fg.type == "rgb" and fg.rgb not in ("00000000", "FFFFFFFF", "FF000000"):
                        return True
                    elif fg.type == "theme":
                        return True
        except Exception:
            pass
    return False


def _parse_xlsx_combined(path: Path):
    """_parse_xlsx + _parse_xlsx_tables를 단일 load_workbook 호출로 통합.

    Returns:
        (parse_result, tables) — parse_result는 parse_excel() 반환 형식,
        tables는 parse_excel_tables() 반환 형식.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ConversionError("openpyxl 패키지가 필요합니다. pip install openpyxl")

    # read_only=False: 스타일(헤더 인식) + merged_cells 모두 접근 가능
    wb = load_workbook(str(path), data_only=True, read_only=False)
    pages = []
    tables = []
    page_num = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not ws.max_row or not ws.max_column:
            continue

        max_col = ws.max_column

        # 헤더 인식 (스타일)
        has_header = _is_header_row(ws, 1, max_col)

        # 병합 셀 수집 (read_only=False이므로 ws.merged_cells 사용 가능)
        merge_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
        skip_cells: set = set()
        for merge_range in ws.merged_cells.ranges:
            min_r = merge_range.min_row - 1
            min_c = merge_range.min_col - 1
            max_r = merge_range.max_row - 1
            max_c = merge_range.max_col - 1
            rowspan = max_r - min_r + 1
            colspan = max_c - min_c + 1
            merge_map[(min_r, min_c)] = (rowspan, colspan)
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    if (r, c) != (min_r, min_c):
                        skip_cells.add((r, c))

        # 데이터 읽기 (한 번)
        raw_rows: List[List[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
            raw_rows.append([_cell_value_to_str(cell.value) for cell in row])

        raw_rows, max_col = _trim_empty_trailing(raw_rows, max_col)
        if not raw_rows or max_col == 0:
            continue

        flat_rows: List[List[str]] = [
            [
                "" if (r_idx, c_idx) in skip_cells
                else (row[c_idx] if c_idx < len(row) else "")
                for c_idx in range(max_col)
            ]
            for r_idx, row in enumerate(raw_rows)
        ]

        md = _build_markdown_table(flat_rows, sheet_name)
        flat_text = "\n".join(c for row in flat_rows for c in row if c)
        first_line = " ".join(flat_rows[0]).strip() if flat_rows else ""
        page_html = _build_html_table(
            raw_rows,
            header_row=True,
            merge_map=merge_map,
            skip_cells=skip_cells,
        )
        table_html = _build_html_table(
            raw_rows,
            header_row=has_header,
            merge_map=merge_map,
            skip_cells=skip_cells,
        )

        # A파트: 페이지 구조
        page_num += 1
        pages.append({
            "page": page_num,
            "sheet_name": sheet_name,
            "text": "",
            "images": [],
            "elements": [
                {
                    "type": "table",
                    "content": page_html,
                    "markdown": md,
                    "sheet": sheet_name,
                    "bbox": None,
                }
            ],
        })

        # B파트: 표 메타
        tables.append({
            "html": table_html,
            "markdown": md,
            "flat_text": flat_text,
            "first_line": first_line,
        })

    wb.close()

    if not pages:
        pages.append({"page": 1, "text": "(빈 Excel 파일)", "images": [], "elements": []})

    print(f"[INFO] Excel 파싱 완료 (combined): {path.name} → {len(pages)}개 시트")
    return {"page_count": len(pages), "pages": pages}, tables


def _parse_xlsx_tables(path: Path) -> List[dict]:
    """openpyxl로 .xlsx 표 추출 (스타일 기반 헤더 인식)"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ConversionError("openpyxl 패키지가 필요합니다. pip install openpyxl")

    wb = load_workbook(str(path), data_only=True, read_only=False)
    tables = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not ws.max_row or not ws.max_column:
            continue

        max_col = ws.max_column

        # 헤더 인식: 1행 스타일 검사
        has_header = _is_header_row(ws, 1, max_col)

        # 병합 셀 수집
        merge_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
        skip_cells: set = set()
        for merge_range in ws.merged_cells.ranges:
            min_r = merge_range.min_row - 1
            min_c = merge_range.min_col - 1
            max_r = merge_range.max_row - 1
            max_c = merge_range.max_col - 1
            rowspan = max_r - min_r + 1
            colspan = max_c - min_c + 1
            merge_map[(min_r, min_c)] = (rowspan, colspan)
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    if (r, c) != (min_r, min_c):
                        skip_cells.add((r, c))

        # 데이터 읽기
        raw_rows: List[List[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col):
            raw_rows.append([_cell_value_to_str(cell.value) for cell in row])

        raw_rows, max_col = _trim_empty_trailing(raw_rows, max_col)
        if not raw_rows or max_col == 0:
            continue

        html = _build_html_table(raw_rows, header_row=has_header,
                                 merge_map=merge_map, skip_cells=skip_cells)

        flat_rows: List[List[str]] = []
        for r_idx, row in enumerate(raw_rows):
            flat_row = []
            for c_idx in range(max_col):
                if c_idx < len(row):
                    flat_row.append("" if (r_idx, c_idx) in skip_cells else row[c_idx])
                else:
                    flat_row.append("")
            flat_rows.append(flat_row)

        md = _build_markdown_table(flat_rows, sheet_name)
        flat_text = "\n".join(c for row in flat_rows for c in row if c)
        first_line = " ".join(flat_rows[0]).strip() if flat_rows else ""

        tables.append({
            "html": html,
            "markdown": md,
            "flat_text": flat_text,
            "first_line": first_line,
        })

    wb.close()
    return tables


def _parse_xls_tables(path: Path) -> List[dict]:
    """xlrd로 .xls 표 추출 (스타일 정보 없으므로 1행은 항상 헤더로 간주)"""
    try:
        import xlrd
    except ImportError:
        # xlrd 없으면 LibreOffice로 xlsx 변환 후 재시도
        import tempfile as _tempfile
        with _tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            _lo = _find_libreoffice()
            if not _lo:
                raise ConversionError(".xls 표 추출에 xlrd 또는 LibreOffice가 필요합니다.")
            cmd = [_lo, "--headless", "--convert-to", "xlsx",
                   "--outdir", str(tmp_path), str(path)]
            try:
                subprocess.run(cmd, capture_output=True, timeout=120, check=True)
            except (subprocess.SubprocessError, OSError) as e:
                raise ConversionError(f".xls 변환 실패: {e}")
            xlsx_path = tmp_path / f"{path.stem}.xlsx"
            if xlsx_path.exists():
                return _parse_xlsx_tables(xlsx_path)
        raise ConversionError(".xls 표 추출 실패")

    wb = xlrd.open_workbook(str(path))
    tables = []

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        sheet_name = ws.name
        if ws.nrows == 0 or ws.ncols == 0:
            continue

        merge_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
        skip_cells: set = set()
        for rlo, rhi, clo, chi in ws.merged_cells:
            rowspan = rhi - rlo
            colspan = chi - clo
            merge_map[(rlo, clo)] = (rowspan, colspan)
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    if (r, c) != (rlo, clo):
                        skip_cells.add((r, c))

        raw_rows: List[List[str]] = []
        for r in range(ws.nrows):
            row_data = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row_data.append(_normalize_excel_text_value(dt))
                    except Exception:
                        row_data.append(str(cell.value))
                else:
                    row_data.append(_cell_value_to_str(cell.value))
            raw_rows.append(row_data)

        raw_rows, max_cols = _trim_empty_trailing(raw_rows, ws.ncols)
        if not raw_rows or max_cols == 0:
            continue

        html = _build_html_table(raw_rows, header_row=True,
                                 merge_map=merge_map, skip_cells=skip_cells)

        flat_rows: List[List[str]] = []
        for r_idx, row in enumerate(raw_rows):
            flat_row = []
            for c_idx in range(max_cols):
                if c_idx < len(row):
                    flat_row.append("" if (r_idx, c_idx) in skip_cells else row[c_idx])
                else:
                    flat_row.append("")
            flat_rows.append(flat_row)

        md = _build_markdown_table(flat_rows, sheet_name)
        flat_text = "\n".join(c for row in flat_rows for c in row if c)
        first_line = " ".join(flat_rows[0]).strip() if flat_rows else ""

        tables.append({
            "html": html,
            "markdown": md,
            "flat_text": flat_text,
            "first_line": first_line,
        })

    wb.release_resources()
    return tables


def parse_excel_tables(path: Path) -> List[dict]:
    """
    Excel 파일에서 표를 추출. parse_hwpx_tables()와 동일한 출력 형식.
    헤더 인식 휴리스틱: 1행에 font.bold 또는 fill.fgColor 있으면 헤더로 간주.

    Returns:
        [{"html": "...", "markdown": "...", "flat_text": "...", "first_line": "..."}, ...]
    """
    ext = path.suffix.lower()
    if ext == ".xls":
        return _parse_xls_tables(path)
    return _parse_xlsx_tables(path)
