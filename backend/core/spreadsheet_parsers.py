from __future__ import annotations

import subprocess
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Tuple

from core.conversion_common import ConversionError, find_libreoffice as _find_libreoffice
from core.spreadsheet_formatting import (
    _build_html_table,
    _build_markdown_table,
    _cell_value_to_str,
    _normalize_excel_text_value,
    _trim_empty_trailing,
)

def parse_excel(path: Path) -> Dict[str, Any]:
    """
    Excel 파일(.xlsx)의 모든 시트를 직접 파싱하여 구조화된 결과 반환.
    VLM 없이 openpyxl로 정확한 데이터 구조를 보존합니다.

    .xls 파일은 xlrd로 시도하고, 실패 시 LibreOffice로 .xlsx 변환 후 재시도합니다.

    Returns:
        build_document_text()와 호환되는 딕셔너리:
        {
            "page_count": int,
            "pages": [
                {
                    "page": int,
                    "text": "",
                    "images": [],
                    "elements": [{"type": "table", "content": html, "markdown": md, "bbox": [...]}],
                },
                ...
            ]
        }
    """
    ext = path.suffix.lower()

    if ext == ".xls":
        return _parse_xls(path)

    return _parse_xlsx(path)


def _xlsx_extract_merge_refs(path: Path) -> Dict[str, List[str]]:
    """xlsx ZIP에서 시트별 병합 셀 ref 목록을 직접 추출 (read_only 모드 보완).

    Returns:
        {"Sheet1": ["A1:C2", "D3:E5"], ...}
    """
    from defusedxml import ElementTree

    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    result: Dict[str, List[str]] = {}

    try:
        with zipfile.ZipFile(path, "r") as zf:
            names = zf.namelist()

            # 시트 이름 순서를 workbook.xml에서 파악
            sheet_names: List[str] = []
            if "xl/workbook.xml" in names:
                with zf.open("xl/workbook.xml") as f:
                    root = ElementTree.parse(f).getroot()
                    for s in root.findall(f".//{{{NS}}}sheet"):
                        sheet_names.append(s.get("name", ""))

            # xl/worksheets/sheet1.xml, sheet2.xml … 순서대로 읽기
            sheet_files = sorted(
                [n for n in names if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
            )
            for i, sf in enumerate(sheet_files):
                sname = sheet_names[i] if i < len(sheet_names) else f"Sheet{i + 1}"
                try:
                    with zf.open(sf) as f:
                        root = ElementTree.parse(f).getroot()
                        refs = [mc.get("ref", "") for mc in root.findall(f".//{{{NS}}}mergeCell")]
                        if refs:
                            result[sname] = refs
                except Exception:
                    pass
    except Exception:
        pass

    return result


def _build_merge_maps_from_refs(
    refs: List[str],
) -> Tuple[Dict[Tuple[int, int], Tuple[int, int]], set]:
    """병합 셀 ref 문자열("A1:C3") → merge_map, skip_cells (0-indexed)"""
    from openpyxl.utils import coordinate_to_tuple

    merge_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
    skip_cells: set = set()

    for ref in refs:
        if ":" not in ref:
            continue
        try:
            tl, br = ref.split(":")
            min_r, min_c = coordinate_to_tuple(tl)
            max_r, max_c = coordinate_to_tuple(br)
            # openpyxl은 1-indexed → 0-indexed 변환
            min_r -= 1
            min_c -= 1
            max_r -= 1
            max_c -= 1
        except Exception:
            continue

        rowspan = max_r - min_r + 1
        colspan = max_c - min_c + 1
        merge_map[(min_r, min_c)] = (rowspan, colspan)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    skip_cells.add((r, c))

    return merge_map, skip_cells


def _parse_xlsx(path: Path) -> Dict[str, Any]:
    """openpyxl을 사용하여 .xlsx 파일 파싱.

    read_only=True + values_only=True 로 데이터를 읽어 속도를 높이고,
    병합 셀 정보는 ZIP XML에서 직접 추출한다.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ConversionError("openpyxl 패키지가 필요합니다. pip install openpyxl")

    # 병합 셀: read_only 모드에서 ws.merged_cells 불가 → XML에서 직접 추출
    merge_info = _xlsx_extract_merge_refs(path)

    # read_only=True: 스타일/수식 로드 생략 → 대용량 파일에서 3~5배 빠름
    wb = load_workbook(str(path), data_only=True, read_only=True)
    pages = []
    page_num = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 병합 셀 맵 구성
        merge_map, skip_cells = _build_merge_maps_from_refs(merge_info.get(sheet_name, []))

        # 데이터 읽기: values_only=True 로 Cell 객체 생성 없이 값만 반환
        raw_rows: List[List[str]] = [
            [_cell_value_to_str(v) for v in row]
            for row in ws.iter_rows(values_only=True)
        ]

        max_cols = max((len(r) for r in raw_rows), default=0)
        raw_rows, max_cols = _trim_empty_trailing(raw_rows, max_cols)

        if not raw_rows or max_cols == 0:
            continue

        # HTML 표 생성
        html = _build_html_table(raw_rows, header_row=True,
                                 merge_map=merge_map, skip_cells=skip_cells)

        # Markdown용 flat_rows (병합 건너뛴 셀은 빈 문자열)
        flat_rows: List[List[str]] = [
            [
                "" if (r_idx, c_idx) in skip_cells
                else (row[c_idx] if c_idx < len(row) else "")
                for c_idx in range(max_cols)
            ]
            for r_idx, row in enumerate(raw_rows)
        ]

        md = _build_markdown_table(flat_rows, sheet_name)

        page_num += 1
        pages.append({
            "page": page_num,
            "sheet_name": sheet_name,
            "text": "",
            "images": [],
            "elements": [
                {
                    "type": "table",
                    "content": html,
                    "markdown": md,
                    "sheet": sheet_name,
                    "bbox": None,
                }
            ],
        })

    wb.close()

    # 빈 파일 처리
    if not pages:
        pages.append({
            "page": 1,
            "text": "(빈 Excel 파일)",
            "images": [],
            "elements": [],
        })

    print(f"[INFO] Excel 파싱 완료: {path.name} → {len(pages)}개 시트")
    return {
        "page_count": len(pages),
        "pages": pages,
    }


def _parse_xls(path: Path) -> Dict[str, Any]:
    """
    .xls 파일 파싱.
    1차: xlrd로 직접 파싱 시도
    2차: LibreOffice로 .xlsx 변환 후 openpyxl로 파싱
    """
    # 1차: xlrd 시도
    try:
        import xlrd  # noqa: F401
        return _parse_xls_with_xlrd(path)
    except ImportError:
        pass
    except Exception as e:
        print(f"[WARNING] xlrd로 .xls 파싱 실패: {e}")

    # 2차: LibreOffice로 .xlsx 변환 후 openpyxl
    print(f"[INFO] .xls → .xlsx 변환 후 파싱 시도: {path.name}")
    import tempfile
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        _lo = _find_libreoffice()
        if _lo:
            cmd = [
                _lo, "--headless", "--convert-to", "xlsx",
                "--outdir", str(tmp_path), str(path),
            ]
            try:
                subprocess.run(cmd, capture_output=True, timeout=120, check=True)
            except (subprocess.SubprocessError, OSError) as e:
                raise ConversionError(f".xls 변환 실패 (LibreOffice): {e}")

            xlsx_path = tmp_path / f"{path.stem}.xlsx"
            if xlsx_path.exists():
                return _parse_xlsx(xlsx_path)

        raise ConversionError(
            ".xls 파일을 처리하려면 xlrd 패키지 또는 LibreOffice가 필요합니다."
        )


def _parse_xls_with_xlrd(path: Path) -> Dict[str, Any]:
    """xlrd를 사용하여 .xls 파일 파싱"""
    import xlrd

    wb = xlrd.open_workbook(str(path))
    pages = []
    page_num = 0

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        sheet_name = ws.name

        if ws.nrows == 0 or ws.ncols == 0:
            continue

        # xlrd는 병합 셀 정보를 merged_cells로 제공
        merge_map = {}
        skip_cells = set()

        for rlo, rhi, clo, chi in ws.merged_cells:
            rowspan = rhi - rlo
            colspan = chi - clo
            merge_map[(rlo, clo)] = (rowspan, colspan)
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    if (r, c) != (rlo, clo):
                        skip_cells.add((r, c))

        raw_rows: List[List[str]] = []
        max_cols = ws.ncols

        for r in range(ws.nrows):
            row_data = []
            for c in range(max_cols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row_data.append(_normalize_excel_text_value(dt))
                    except Exception:
                        row_data.append(str(cell.value))
                else:
                    row_data.append(_cell_value_to_str(cell.value))
            raw_rows.append(row_data)

        raw_rows, max_cols = _trim_empty_trailing(raw_rows, max_cols)

        if not raw_rows or max_cols == 0:
            continue

        html = _build_html_table(raw_rows, header_row=True,
                                 merge_map=merge_map, skip_cells=skip_cells)

        flat_rows: List[List[str]] = [
            [
                "" if (r_idx, c_idx) in skip_cells
                else (row[c_idx] if c_idx < len(row) else "")
                for c_idx in range(max_cols)
            ]
            for r_idx, row in enumerate(raw_rows)
        ]

        md = _build_markdown_table(flat_rows, sheet_name)

        page_num += 1
        pages.append({
            "page": page_num,
            "sheet_name": sheet_name,
            "text": "",
            "images": [],
            "elements": [
                {
                    "type": "table",
                    "content": html,
                    "markdown": md,
                    "sheet": sheet_name,
                    "bbox": None,
                }
            ],
        })

    wb.release_resources()

    if not pages:
        pages.append({
            "page": 1,
            "text": "(빈 Excel 파일)",
            "images": [],
            "elements": [],
        })

    print(f"[INFO] Excel(.xls) 파싱 완료: {path.name} → {len(pages)}개 시트")
    return {
        "page_count": len(pages),
        "pages": pages,
    }
