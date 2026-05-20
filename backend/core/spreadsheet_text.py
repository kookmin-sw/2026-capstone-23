from __future__ import annotations

from pathlib import Path
from typing import Any

from core.conversion_common import ConversionError
from core.spreadsheet_formatting import _normalize_excel_text_value

def _extract_text_from_xlsx_cell(cell: Any) -> str:
    """openpyxl 셀 타입을 기준으로 텍스트 추출용 문자열 변환"""
    if cell.value is None:
        return ""

    # openpyxl data_type: n=숫자, s=문자, d=날짜, b=bool
    if cell.is_date or cell.data_type == "d":
        return _normalize_excel_text_value(cell.value)
    if cell.data_type == "b":
        return "TRUE" if bool(cell.value) else "FALSE"
    if cell.data_type == "n":
        return _normalize_excel_text_value(cell.value)
    if cell.data_type in {"s", "str", "inlineStr"}:
        return str(cell.value).strip()

    # 수식/에러 등 기타 타입은 현재 값을 그대로 텍스트화
    return _normalize_excel_text_value(cell.value)


def _extract_text_from_xls_cell(cell: Any, datemode: int) -> str:
    """xlrd 셀 타입을 기준으로 텍스트 추출용 문자열 변환"""
    import xlrd

    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return _normalize_excel_text_value(xlrd.xldate_as_datetime(cell.value, datemode))
        except Exception:
            return _normalize_excel_text_value(cell.value)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if bool(cell.value) else "FALSE"
    if cell.ctype in {xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_TEXT}:
        return _normalize_excel_text_value(cell.value)

    # 공식, 에러 등 기타 타입은 현재 값을 그대로 텍스트화
    return _normalize_excel_text_value(cell.value)


def _extract_text_from_xlsx(path: Path) -> str:
    """
    .xlsx 파일에서 시트별 텍스트를 추출하여 하나의 문자열로 반환
    반환 예:
    Sheet1
    값1 값2
    값3 값4

    Sheet2
    ...
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ConversionError("openpyxl 패키지가 필요합니다. pip install openpyxl")

    if not path.exists():
        raise ConversionError(f"Excel 파일을 찾을 수 없습니다: {path}")
    if path.stat().st_size == 0:
        raise ConversionError(f"Excel 파일이 비어있습니다: {path}")

    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
        sheet_blocks: list[str] = []
        try:
            for ws in wb.worksheets:
                lines = [ws.title]
                has_content = False

                for row in ws.iter_rows():
                    row_values = []
                    for cell in row:
                        text = _extract_text_from_xlsx_cell(cell)
                        if text:
                            row_values.append(text)

                    if row_values:
                        lines.append(" ".join(row_values))
                        has_content = True

                if has_content:
                    sheet_blocks.append("\n".join(lines))
        finally:
            wb.close()

        if sheet_blocks:
            return "\n\n".join(sheet_blocks)
        raise ConversionError("Excel(.xlsx)에서 텍스트를 추출할 수 없습니다.")
    except Exception as e:
        if isinstance(e, ConversionError):
            raise
        raise ConversionError(f"Excel(.xlsx) 텍스트 추출 실패: {e}")


def _extract_text_from_xls(path: Path) -> str:
    """
    .xls 파일에서 시트별 텍스트를 추출하여 하나의 문자열로 반환
    xlrd 셀 타입을 사용해 기본 타입별 문자열 변환을 적용합니다.
    """
    try:
        import xlrd
    except ImportError:
        raise ConversionError("xlrd 패키지가 필요합니다. pip install xlrd")

    if not path.exists():
        raise ConversionError(f"Excel 파일을 찾을 수 없습니다: {path}")
    if path.stat().st_size == 0:
        raise ConversionError(f"Excel 파일이 비어있습니다: {path}")

    try:
        wb = xlrd.open_workbook(str(path))
        sheet_blocks: list[str] = []
        try:
            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                lines = [ws.name]
                has_content = False

                for row_idx in range(ws.nrows):
                    row_values = []
                    for col_idx in range(ws.ncols):
                        cell = ws.cell(row_idx, col_idx)
                        text = _extract_text_from_xls_cell(cell, wb.datemode)
                        if text:
                            row_values.append(text)

                    if row_values:
                        lines.append(" ".join(row_values))
                        has_content = True

                if has_content:
                    sheet_blocks.append("\n".join(lines))
        finally:
            wb.release_resources()

        if sheet_blocks:
            return "\n\n".join(sheet_blocks)
        raise ConversionError("Excel(.xls)에서 텍스트를 추출할 수 없습니다.")
    except Exception as e:
        if isinstance(e, ConversionError):
            raise
        raise ConversionError(f"Excel(.xls) 텍스트 추출 실패: {e}")


def extract_text_from_excel(path: Path) -> str:
    """엑셀 파일(.xlsx, .xls)에서 시트별 텍스트를 추출"""
    ext = path.suffix.lower()

    if ext == ".xlsx":
        return _extract_text_from_xlsx(path)
    if ext == ".xls":
        return _extract_text_from_xls(path)

    raise ConversionError(f"지원하지 않는 Excel 형식입니다: {path.suffix}")
