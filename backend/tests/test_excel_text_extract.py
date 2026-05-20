from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from core.converters import extract_text_from_excel


def test_extract_text_from_xlsx(tmp_path: Path):
    sample = tmp_path / "sample.xlsx"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "값1"
    ws1["B1"] = "값2"
    ws1["A2"] = 100
    ws1["B2"] = True

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "일자"
    ws2["B1"] = "설명"
    ws2["A2"] = datetime(2026, 3, 25, 9, 30, 0)
    ws2["B2"] = "회의"
    wb.save(sample)
    wb.close()

    assert extract_text_from_excel(sample) == (
        "Sheet1\n"
        "값1 값2\n"
        "100 TRUE\n\n"
        "Sheet2\n"
        "일자 설명\n"
        "2026-03-25 09:30:00 회의"
    )


def test_extract_text_from_xls_with_xlrd_stub(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    sample = tmp_path / "sample.xls"
    sample.write_bytes(b"stub")

    class FakeSheet:
        def __init__(self, name, rows):
            self.name = name
            self._rows = rows
            self.nrows = len(rows)
            self.ncols = len(rows[0]) if rows else 0

        def cell(self, row_idx, col_idx):
            return self._rows[row_idx][col_idx]

    class FakeWorkbook:
        def __init__(self):
            self.datemode = 0
            self._sheets = [
                FakeSheet(
                    "Sheet1",
                    [
                        [SimpleNamespace(ctype=1, value="값1"), SimpleNamespace(ctype=1, value="값2")],
                        [SimpleNamespace(ctype=2, value=10.0), SimpleNamespace(ctype=4, value=1)],
                    ],
                ),
                FakeSheet(
                    "Sheet2",
                    [
                        [SimpleNamespace(ctype=1, value="메모")],
                    ],
                ),
            ]
            self.nsheets = len(self._sheets)

        def sheet_by_index(self, idx):
            return self._sheets[idx]

        def release_resources(self):
            pass

    fake_xlrd = SimpleNamespace(
        XL_CELL_EMPTY=0,
        XL_CELL_TEXT=1,
        XL_CELL_NUMBER=2,
        XL_CELL_DATE=3,
        XL_CELL_BOOLEAN=4,
        XL_CELL_BLANK=6,
        open_workbook=lambda _: FakeWorkbook(),
        xldate_as_datetime=lambda value, datemode: datetime(2026, 3, 25, 0, 0, 0),
    )

    import sys

    monkeypatch.setitem(sys.modules, "xlrd", fake_xlrd)

    assert extract_text_from_excel(sample) == (
        "Sheet1\n"
        "값1 값2\n"
        "10 TRUE\n\n"
        "Sheet2\n"
        "메모"
    )
